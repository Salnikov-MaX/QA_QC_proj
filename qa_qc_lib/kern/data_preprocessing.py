import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from qa_qc_kern import QA_QC_kern


class DataPreprocessing:
    def __init__(self, files=None, glossary_of_names=None):
        if glossary_of_names is None:
            glossary_of_names = {}
        if files is None:
            files = []
        self.wb = Workbook()
        self.ws = self.wb.active
        self.user_glossary_of_names = glossary_of_names
        self.input_files = files
        self.failed_tests = {}
        self.glossary_of_names = {
            "Направление": "",
            "Лабораторный номер": "",
            "Кровля интервала отбора": "",
            "Подошва интервала отбора": "",
            "Место отбора (ниже кровли), м": "",
            "Глубина отбора, м": "",
            "Вынос керна, м": "",
            "Вынос керна, %": "",
            "Ск %": "",
            "Открытая пористость по жидкости": "",
            "Открытая пористость по газу": "",
            "Открытая пористость в пластовых условиях": "",
            "Открытая пористость по керосину": "",
            "Кпр_газ(гелий)": "",
            "Параметр пористости": "",
            "So": "",
            "Кво": "",
            "Плотность абсолютно сухого образца": "",
            "Рн": "",
            "Sw": "",
            "Газопроницаемость по Кликенбергу": "",
            "Газопроницаемость, mkm2 (parallel)": "",
            "Газопроницаемость Кликенбергу": "",
            "Объемная плотность": "",
            "Минералогическая плотность": "",
            "Эффективная проницаемость": "",
            "Ск": "",
            "Параметр насыщения": "",
            "Газопроницаемость по воде": "",
            "Плотность максимально увлажненного образца": "",
            "Упругие свойства": "",
            "Примечание": ""
        }
        self.data_dict = {
            "Направление": None,
            "Лабораторный номер": None,
            "Кровля интервала отбора": None,
            "Подошва интервала отбора": None,
            "Место отбора (ниже кровли), м": None,
            "Глубина отбора, м": None,
            "Вынос керна, м": None,
            "Вынос керна, %": None,
            "Ск %": None,
            "Открытая пористость по жидкости": None,
            "Открытая пористость по газу": None,
            "Открытая пористость в пластовых условиях": None,
            "Открытая пористость по керосину": None,
            "Кпр_газ(гелий)": None,
            "Параметр пористости": None,
            "Эффективная проницаемость": None,
            "So": None,
            "Газопроницаемость по Кликенбергу": None,
            "Кво": None,
            "Параметр насыщения": None,
            "Плотность абсолютно сухого образца": None,
            "Рн": None,
            "Sw": None,
            "Газопроницаемость, mkm2 (parallel)": None,
            "Газопроницаемость Кликенбергу": None,
            "Объемная плотность": None,
            "Минералогическая плотность": None,
            "Ск": None,
            "Газопроницаемость по воде": None,
            "Плотность максимально увлажненного образца": None,
            "Упругие свойства": None,
            "Примечание": None
        }
        self.__create_new_glossary()
        self.headers = [
            "Лабораторный номер", "Кровля интервала отбора", "Подошва интервала отбора",
            "Эффективная проницаемость", "Параметр насыщения",
            "Место отбора (ниже кровли), м", "Глубина отбора, м",
            "Вынос керна, м", "Вынос керна, %", "Ск %", "Открытая пористость по жидкости",
            "Открытая пористость по газу",
            "Открытая пористость в пластовых условиях", "Открытая пористость по керосину", "Кпр_газ(гелий)",
            "Параметр пористости", "So",
            "Кво", "Плотность абсолютно сухого образца", "Рн",
            "Sw", "Газопроницаемость, mkm2 (parallel)", "Газопроницаемость Кликенбергу",
            "Объемная плотность", "Минералогическая плотность", "Ск", "Газопроницаемость по воде",
            "Плотность максимально увлажненного образца"
            "Упругие свойства",
            "Примечание", "Направление"
        ]
        self.parallel_density = []
        self.parallel_porosity = []
        self.parallel_number = []
        self.perpendicular_density = []
        self.perpendicular_porosity = []
        self.perpendicular_number = []
        self.parallel_carbonate = []
        self.interval = []
        self.perpendicular_carbonate = []

    def __create_new_glossary(self):
        for key in self.user_glossary_of_names:
            if key in self.glossary_of_names:
                self.glossary_of_names[key] = self.user_glossary_of_names[key]

    def process_files(self):
        for idx, header in enumerate(self.headers, start=1):
            self.ws.cell(row=2, column=idx, value=header)
            if len(header) > 7:
                letter = get_column_letter(idx)
                self.ws.column_dimensions[letter].width = 20
        df = []
        for file in self.input_files:
            file_name = os.path.basename(file)
            if file.endswith(".xlsx") or file.endswith(".xls"):
                df = pd.read_excel(file)
            elif file.endswith(".txt"):
                df = pd.read_table(file, sep='\t', header=0, on_bad_lines='skip')
            if len(df) != 0:
                for col in df.columns:
                    if col in self.glossary_of_names.values():
                        key = list(self.glossary_of_names.keys())[list(self.glossary_of_names.values()).index(col)]
                        col_data = df[col].tolist()
                        processed_col_data = []
                        for value in col_data:
                            try:
                                processed_value = float(value.replace(',', '.'))
                            except:
                                processed_value = value
                            processed_col_data.append(processed_value)
                        self.data_dict[key] = processed_col_data
                        processed_depth = []
                        try:
                            for value in df[self.user_glossary_of_names[file_name]]:
                                try:
                                    processed_value = float(value.replace(',', '.'))
                                except:
                                    processed_value = value
                                processed_depth.append(processed_value)
                            new_depths = processed_depth
                        except KeyError:
                            raise ValueError(f"Не указана глубина для файла {file_name}")

                        existing_depths = self.data_dict["Глубина отбора, м"]
                        if existing_depths is None:
                            self.data_dict["Глубина отбора, м"] = new_depths
                            existing_depths = []

                        for existing_depth, new_depth in zip(existing_depths, new_depths):
                            if abs(existing_depth - new_depth) > 0.1:
                                print("Depth values in the file are not consistent")

                        if len(existing_depths) != len(new_depths) and len(existing_depths) != 0:
                            raise ValueError("Разные глубины")
                        for idx, (key, values) in enumerate(self.data_dict.items(), start=2):
                            if values is not None and key != "Глубина отбора, м":
                                for row_idx, value in enumerate(values, start=2):
                                    col_idx = self.headers.index(key) + 1
                                    depth_value = self.data_dict["Глубина отбора, м"][row_idx - 2]
                                    if depth_value == new_depths[row_idx - 2]:
                                        self.ws.cell(row=row_idx + 1, column=col_idx, value=value)

                        for row_idx, value in enumerate(self.data_dict["Глубина отбора, м"], start=2):
                            col_idx = self.headers.index("Глубина отбора, м") + 1
                            self.ws.cell(row=row_idx + 1, column=col_idx, value=value)

    def start_tests(self, tests_name=None):
        if tests_name is None:
            tests_name = []
        if self.data_dict["Направление"] is not None and \
                self.data_dict["Ск"] is not None and \
                self.data_dict["Лабораторный номер"] is not None and \
                self.data_dict["Открытая пористость по жидкости"] is not None and \
                self.data_dict["Плотность абсолютно сухого образца"]:
            self.parallel_data_parsing()
        if self.data_dict["Кровля интервала отбора"] is not None and self.data_dict["Подошва интервала отбора"]:
            self.interval_data_parsing()
        test_system = QA_QC_kern(pas=np.array(self.data_dict["Плотность абсолютно сухого образца"]),
                                 note=np.array(self.data_dict["Примечание"]),
                                 kno=np.array(self.data_dict["So"]),
                                 kp_plast=np.array(self.data_dict["Открытая пористость в пластовых условиях"]),
                                 density=np.array(self.data_dict["Плотность абсолютно сухого образца"]),
                                 water_permeability=np.array(self.data_dict["Газопроницаемость по воде"]),
                                 kp_pov=np.array(self.data_dict["Открытая пористость по жидкости"]),
                                 perpendicular=np.array(self.perpendicular_number),
                                 perpendicular_porosity=np.array(self.perpendicular_porosity),
                                 perpendicular_density=np.array(self.perpendicular_density),
                                 perpendicular_carbonate=np.array(self.perpendicular_carbonate),
                                 parallel=np.array(self.parallel_number),
                                 parallel_porosity=np.array(self.parallel_porosity),
                                 parallel_density=np.array(self.parallel_density),
                                 parallel_carbonate=np.array(self.parallel_carbonate),
                                 kp=np.array(self.data_dict["Открытая пористость по жидкости"]),
                                 top=np.array(self.data_dict["Кровля интервала отбора"]),
                                 core_removal_in_meters=np.array(self.data_dict["Вынос керна, м"]),
                                 intervals=self.interval,
                                 bottom=np.array(self.data_dict["Подошва интервала отбора"]),
                                 percent_core_removal=np.array(self.data_dict["Вынос керна, %"]),
                                 outreach_in_meters=np.array(self.data_dict["Вынос керна, м"]),
                                 sw_residual=np.array(self.data_dict["Кво"]),
                                 core_sampling=np.array(self.data_dict["Глубина отбора, м"]),
                                 kpr=np.array(self.data_dict["Кпр_газ(гелий)"]),
                                 rp=np.array(self.data_dict["Параметр пористости"]),
                                 pmu=np.array(self.data_dict["Плотность максимально увлажненного образца"]),
                                 rn=np.array(self.data_dict["Параметр насыщения"]),
                                 obplnas=np.array(self.data_dict["Плотность абсолютно сухого образца"]),
                                 poroTBU=np.array(self.data_dict["Открытая пористость в пластовых условиях"]),
                                 poroHe=np.array(self.data_dict["Открытая пористость по газу"]),
                                 porosity_open=np.array(self.data_dict["Открытая пористость по жидкости"]),
                                 porosity_kerosine=np.array(self.data_dict["Открытая пористость по керосину"]),
                                 sw=np.array(self.data_dict["Sw"]),
                                 parallel_permeability=np.array(self.data_dict["Газопроницаемость, mkm2 (parallel)"]),
                                 klickenberg_permeability=np.array(self.data_dict["Газопроницаемость по Кликенбергу"]),
                                 effective_permeability=np.array(self.data_dict["Эффективная проницаемость"]),
                                 md=np.array(self.data_dict["Место отбора (ниже кровли), м"]),show=False)

        self.failed_tests = test_system.start_tests(tests_name)["wrong_parameters"]
        test_system.generate_test_report()
        self.error_flagging()
        return self.failed_tests

    def parallel_data_parsing(self):
        direction_array = self.data_dict["Направление"]
        for idx, is_parallel in enumerate(direction_array):
            if is_parallel == 1:
                self.parallel_density.append([self.data_dict["Плотность абсолютно сухого образца"][idx], idx])
                self.parallel_porosity.append([self.data_dict["Открытая пористость по жидкости"][idx], idx])
                self.parallel_number.append([self.data_dict["Лабораторный номер"][idx], idx])
                self.parallel_carbonate.append([self.data_dict["Ск"][idx], idx])
            else:
                self.perpendicular_density.append([self.data_dict["Плотность абсолютно сухого образца"][idx], idx])
                self.perpendicular_porosity.append([self.data_dict["Открытая пористость по жидкости"][idx], idx])
                self.perpendicular_number.append([self.data_dict["Лабораторный номер"][idx], idx])
                self.perpendicular_carbonate.append([self.data_dict["Ск"][idx], idx])

    def interval_data_parsing(self):
        top = self.data_dict["Кровля интервала отбора"]
        bottom = self.data_dict["Подошва интервала отбора"]
        for i in range(len(top)):
            self.interval.append([top[i], bottom[i]])

    def error_flagging(self):
        self.ws.cell(row=1, column=38, value="Примечание")
        self.ws.merge_cells(start_row=1, start_column=38, end_row=1,
                            end_column=42)
        self.ws.cell(row=1, column=38).alignment = Alignment(horizontal="center",
                                                             vertical="center")
        for idx1, (test_name, failed_columns) in enumerate(self.failed_tests.items(), start=2):
            self.ws.cell(row=2, column=38 + idx1 + 1, value=test_name)
            letter = get_column_letter(38 + idx1 + 1)
            self.ws.column_dimensions[letter].width = 30
            for idx, column_data in enumerate(failed_columns[:-1]):
                for col, values in column_data.items():
                    col_idx = self.headers.index(col) + 1  # Index of the column in the headers
                    for row_idx, value in enumerate(values, start=2):
                        cell = self.ws.cell(row=value + 3, column=col_idx)
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        self.ws.cell(row=int(value) + 3, column=38 + idx1 + 1, value=failed_columns[-1])

    def save_file(self, output_excel_path="kern\\data", file_name="result"):
        self.wb.save(f"{output_excel_path}\\{file_name}.xlsx")


input_files = [
    "kern\\data\\Poro.txt",
    "kern\\data\\59PObraz.xlsx",
    "kern\\data\\Direction.xlsx"
]
dic = {
    "Ск": "Carbonate",
    "Эффективная проницаемость": "PermEf",
    "59PObraz.xlsx": "MD",
    "Direction.xlsx": "Глубина",
    'Плотность абсолютно сухого образца': "Плотность",
    "Параметр пористости": "RI",
    "Направление": "Направление",
    "Лабораторный номер": "Ном",
    "Открытая пористость по жидкости": "Porosity (open)",
    "Открытая пористость по газу": "PoroHe",
    "Открытая пористость в пластовых условиях": "Porosity (open)",
    "Открытая пористость по керосину": "Porosity (kerosine)",
    # "Кпр_газ(гелий)": "Permeability (perpendicular)",
    # "Параметр пористости": "So",
    "So": "So",
    "Poro.txt": "Poro",
    # "Кво": "Sw",
    # "Плотность абсолютно сухого образца": "Density, g/cc",
    # "Рн": "PoroTBU",
    "Sw": "Sw",
    # "Газопроницаемость по Кликенбергу": "PoroTBU",
    "Подошва интервала отбора": "Bottom",
    "Кровля интервала отбора": "Top",
    "Вынос керна, м": "Vynos, m",
    "Газопроницаемость, mkm2 (parallel)": "Permeability, mkm2 (parallel)",
    "Вынос керна, %": "Vynos, %",
    # "Газопроницаемость Кликенбергу": "PoroTBU",
    # "Объемная плотность": "PoroTBU",
    # "Минералогическая плотность": "PoroTBU",
    # "Параметр насыщения": "PoroTBU",
    # "Газопроницаемость по воде": "PoroTBU",
    # "Плотность максимально увлажненного образца": "PoroTBU",
    # "Упругие свойства": "PoroTBU",
    "Примечание": "Примечание"
}

file_modal = DataPreprocessing(files=input_files, glossary_of_names=dic)
file_modal.process_files()
file_modal.start_tests(
    ["test_correctness_of_p_sk_kp", "test_open_porosity", "test_porosity_TBU", "test_porosity_kerosine",
     "test_residual_water_saturation", "test_parallel_permeability", "test_monotony", "test_porosity_HE",
     "test_quo_kp_dependence", "test_kp_density_dependence", "test_effective_permeability", "test_coring_depths_third",
     "test_table_notes", "test_pmu_kp_dependence", "test_quo_and_qno"])
file_modal.save_file()
